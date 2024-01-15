import openpyxl
from user import User
from post import Post

#Requirement to print report in {Supplier_name_1:total_inventory, Supplier_name_2:total_inventory,Supplier_name_3:total_inventory}
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

#print(product_list.max_row)
products_per_supplier = {}
total_value_per_supplier = {}
total_inventory = 0
product_under_10_inventory = {}



for product_row in range(2,product_list.max_row+1):
    product_number = product_list.cell(product_row,1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = product_list.cell(product_row, 5)

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        #  print("adding new Supplier")
        products_per_supplier[supplier_name] = 1

    # calculation total vlue per supplier wise
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price


        #print(supplier_name)
        #print("Current row inventory Value :", inventory)
        #print("Current row Price Value :", float(price))
        #print("Current row Total Value :",total_value_per_supplier[supplier_name])

    else:
        total_value_per_supplier[supplier_name] = inventory + price

    #Logic product  with Inventory less than 10
    if int(inventory) < 10:
        product_under_10_inventory[product_number] = inventory

    inventory_price.value = inventory * price

    #print(product_row, inventory_price.value)

#print(products_per_supplier)
#print(total_value_per_supplier)
#print(product_under_10_inventory)

inv_file.save("inventory_with_total_value.xlsx")

app_user_one = User("santoshkumar78@gmail.com", "Santosh Kumar", "pass@123", "Mananger")
app_user_one.get_user_info()

app_user_two = User("sanket@gmail.com", "Sanket", "pas11@123", "Mananger")
app_user_two.get_user_info()

new_post = Post("This is my life", app_user_one.name)
new_post.get_post_info()




